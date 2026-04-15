# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import os
import io

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Column configurations and data types
COLUMNS_CONFIG = [
    {'name': 'No.', 'type': 'number', 'required': True},
    {'name': 'Full Name', 'type': 'text', 'required': True},
    {'name': 'Gender', 'type': 'select', 'options': ['Male', 'Female'], 'required': True},
    {'name': 'Birth Year', 'type': 'number', 'required': False},
    {'name': 'Date of Birth', 'type': 'date', 'required': False},
    {'name': 'Place of Birth', 'type': 'text', 'required': False},
    {'name': 'Father Name', 'type': 'text', 'required': False},
    {'name': 'Mother Name', 'type': 'text', 'required': False},
    {'name': 'Address', 'type': 'text', 'required': False},
    {'name': 'Phone Number', 'type': 'text', 'required': False},
    {'name': 'Email', 'type': 'email', 'required': False},
    {'name': 'Notes', 'type': 'text', 'required': False},
]

def get_column_names():
    """Get list of column names"""
    return [col['name'] for col in COLUMNS_CONFIG]

def create_excel_file(data, filename=None):
    """Generate Excel file from data"""
    if filename is None:
        filename = f"data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Header formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Apply header formatting
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Apply data formatting
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                col_idx = cell.column - 1
                if col_idx < len(COLUMNS_CONFIG):
                    col_config = COLUMNS_CONFIG[col_idx]
                    if col_config['type'] == 'number':
                        cell.number_format = '0'
                    elif col_config['type'] == 'date':
                        cell.number_format = 'dd/mm/yyyy'

        # Adjust column widths based on English names
        for i, col_config in enumerate(COLUMNS_CONFIG, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            if col_config['name'] in ['Address', 'Notes', 'Email']:
                worksheet.column_dimensions[col_letter].width = 30
            elif col_config['name'] in ['Father Name', 'Mother Name', 'Place of Birth', 'Full Name']:
                worksheet.column_dimensions[col_letter].width = 20
            else:
                worksheet.column_dimensions[col_letter].width = 15

        worksheet.row_dimensions[1].height = 25

    output.seek(0)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    with open(filepath, 'wb') as f:
        f.write(output.getvalue())

    return filepath, filename

def parse_excel_file(filepath):
    """Read Excel file and return data as dictionary list"""
    df = pd.read_excel(filepath)
    df = df.fillna('')
    data = []
    
    for idx, row in df.iterrows():
        row_data = {}
        for col in get_column_names():
            if col in df.columns:
                val = row[col]
                if col in ['Date of Birth'] and pd.notna(val):
                    try:
                        date_val = pd.to_datetime(val)
                        row_data[col] = date_val.strftime('%Y-%m-%d')
                    except:
                        row_data[col] = str(val)
                else:
                    row_data[col] = str(val) if pd.notna(val) else ''
            else:
                row_data[col] = ''
        data.append(row_data)

    return data

@app.route('/')
def index():
    """Main application route"""
    return render_template('index.html', columns=COLUMNS_CONFIG)

@app.route('/api/columns', methods=['GET'])
def get_columns():
    """API to get column configurations"""
    return jsonify(COLUMNS_CONFIG)

@app.route('/api/data', methods=['GET'])
def get_data():
    """API to fetch existing data"""
    return jsonify([])

@app.route('/api/data', methods=['POST'])
def save_data():
    """API to save new data"""
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400

    try:
        # Validate required fields
        for row in data:
            for col_config in COLUMNS_CONFIG:
                col_name = col_config['name']
                if col_config['required'] and not str(row.get(col_name, '')).strip():
                    return jsonify({
                        'success': False,
                        'error': f'Column "{col_name}" is required.'
                    }), 400

        filepath, filename = create_excel_file(data)
        return jsonify({
            'success': True,
            'message': 'Data saved successfully',
            'filename': filename
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download/<filename>')
def download_file(filename):
    """API to download generated Excel file"""
    try:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(filepath):
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """API to upload and parse an existing Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed'}), 400

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        data = parse_excel_file(filepath)

        return jsonify({
            'success': True,
            'data': data,
            'message': f'Successfully uploaded {len(data)} rows.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def export_data():
    """API to trigger data export"""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided to export'}), 400

        filepath, filename = create_excel_file(data)
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/api/download/{filename}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)